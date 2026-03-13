# -*- coding: utf-8 -*-
"""Parla programma deadline – professional XLSX döredýär. Python 3.8+."""
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl gerek: pip install openpyxl")
    raise

DIR = Path(__file__).resolve().parent
OUTPUT_HOME = DIR / "PARLA_HOME_DEADLINE.xlsx"
OUTPUT_MAIN = DIR / "PARLA_PROGRAMMA_DEADLINE.xlsx"

# Diňe Home page (Sahypa) üçin – has detalny deadline.
# Sütunlar: №, Wepaly, Deadline, Bölüm (Sahypa bölümleri), Dereje, Ýagdaý, Bellik
HEADERS = ["№", "Wepaly", "Deadline", "Bölüm", "Dereje", "Ýagdaý", "Bellik"]

ROWS = [
    ["Header: Gözle başlygy, ýerleşiş chip (tap → location picker), bildiriş ikony we badge", "2026-03-12", "Header", "Ýokary", "Garaşylýar", ""],
    ["Header: profil initial (ad ilkinji harpy), tap → Profil tab", "2026-03-12", "Header", "Ýokary", "Garaşylýar", ""],
    ["Pill gözleg: search bar, tap → gözleg ekrany, semantics we haptic", "2026-03-12", "Gözleg", "Ýokary", "Garaşylýar", ""],
    ["Kategoriýalar: chip listi, scroll sag fade (gradient), tap → salonlar listi/kategoriýa", "2026-03-14", "Kategoriýalar", "Ýokary", "Garaşylýar", ""],
    ["Arzanladyşlar: gradient fon, badge (mysal: 2 gün), tap → deals ekrany", "2026-03-14", "Arzanladyşlar", "Orta", "Garaşylýar", ""],
    ["Salon kart (home): Material/InkWell, surat, ad, ýer, favourite ýürek (animasiýa)", "2026-03-15", "Salon kart", "Ýokary", "Garaşylýar", ""],
    ["Salon kart: semantics (Semantics/Label), haptic basanda", "2026-03-15", "Salon kart", "Orta", "Garaşylýar", ""],
    ["Boş sanaw: salon ýok ýagdaýy, mesaj we Sahypa täzelemek / ýerleşiş üýtget", "2026-03-15", "Sanaw", "Ýokary", "Garaşylýar", ""],
    ["Ýerleşiş filtiri: saýlanan ýere görä salonlar süzülýär (chip bilen birikeli)", "2026-03-15", "Ýerleşiş", "Ýokary", "Garaşylýar", ""],
    ["RefreshIndicator: çekip täzelemek, loading reňk tema bilen", "2026-03-16", "Umumy", "Orta", "Garaşylýar", ""],
    ["Loading/error: salonlar ýüklenýärkä görkeziş, ýalňyşlyk mesajy", "2026-03-16", "Umumy", "Orta", "Garaşylýar", ""],
    # ── Fresha vs Parla analizi: Home page kemçilikleri (FRESHA_VS_PARLA_UI_ANALIZ.md) ──
    ["Salon kart: reýting/syn backend-den ýa-da aýdyň mock belgisi (Fresha-da hakyky)", "2026-03-20", "Salon kart", "Ýokary", "Garaşylýar", "Kemçilik 1"],
    ["Salon kart: uzaklyk GPS/ýerleşiş bilen ýa-da mock belgisi (Fresha near me)", "2026-03-20", "Salon kart", "Ýokary", "Garaşylýar", "Kemçilik 2"],
    ["Bölümler Maslahat berilýän/Meşhur/Täze: aýry data ýa-da aýdyň görkeziş (Fresha curation)", "2026-03-22", "Sahypa", "Orta", "Garaşylýar", "Kemçilik 3"],
    ["Salon kartda Açyk/Ýapyk ýa-da iş wagty (backend taýýar bolsa) (Fresha open now)", "2026-03-22", "Salon kart", "Orta", "Garaşylýar", "Kemçilik 4"],
    ["Arzanladyşlar: backend API ýa-da Täzelikler ýakyn mesajy (Fresha deals)", "2026-03-22", "Arzanladyşlar", "Orta", "Garaşylýar", "Kemçilik 5"],
    ["Kategoriýa scroll: çep tarapda fade gradient (Fresha meňzeş)", "2026-03-22", "Kategoriýalar", "Pes", "Garaşylýar", "Kemçilik 6"],
    ["Salon kartda çalt Bron et düwmäsi (opsional) (Fresha quick book)", "2026-03-25", "Salon kart", "Pes", "Garaşylýar", "Kemçilik 7"],
    ["Boş netije (ýerleşiş): Ýerleşişi üýtget düwmäsi aýdyň (Fresha empty state)", "2026-03-25", "Sanaw", "Orta", "Garaşylýar", "Kemçilik 8"],
    ["Loading: shimmer efekt ýa-da skeleton täzelemek (Fresha style)", "2026-03-25", "Umumy", "Pes", "Garaşylýar", "Kemçilik 9"],
    ["Section semantics: Maslahat berilýän we ş.m. bölümler üçin accessibility", "2026-03-28", "Umumy", "Orta", "Garaşylýar", "Kemçilik 10"],
]

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Işler"

    # ---- Stiller ----
    header_fill = PatternFill(start_color="00ACC1", end_color="00ACC1", fill_type="solid")  # Parla cyan
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- Sahypa 1: Işler ----
    # Başlyk (A1)
    ws.merge_cells("A1:G1")
    ws["A1"] = "Parla — Home page (Sahypa) deadline – diňe Sahypa"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Info setir (A2)
    ws.merge_cells("A2:G2")
    ws["A2"] = "Diňe Home page (Sahypa): header, gözleg, kategoriýalar, arzanladyşlar, salon kartlar, boş sanaw, ýerleşiş. Ýagdaý: Garaşylýar / Tamamlandy."
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws.row_dimensions[2].height = 18

    # Header (3-nji setir)
    for col, title in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align
    ws.row_dimensions[3].height = 22

    # Maglumat setirleri
    for i, row_data in enumerate(ROWS, 1):
        row_idx = 3 + i
        ws.cell(row=row_idx, column=1, value=i)
        ws.cell(row=row_idx, column=2, value=row_data[0])  # Wepaly
        # Deadline – sene formaty
        try:
            dt = datetime.strptime(row_data[1], "%Y-%m-%d")
            ws.cell(row=row_idx, column=3, value=dt)
            ws.cell(row=row_idx, column=3).number_format = "YYYY-MM-DD"
        except Exception:
            ws.cell(row=row_idx, column=3, value=row_data[1])
        ws.cell(row=row_idx, column=4, value=row_data[2])  # Bölüm
        ws.cell(row=row_idx, column=5, value=row_data[3])  # Dereje
        ws.cell(row=row_idx, column=6, value=row_data[4] or "Garaşylýar")  # Ýagdaý
        ws.cell(row=row_idx, column=7, value=row_data[5] or "")  # Bellik
        for c in range(1, 8):
            ws.cell(row=row_idx, column=c).border = thin_border
            ws.cell(row=row_idx, column=c).alignment = center_align if c in (1, 3, 4, 5, 6) else left_align

    last_row = 3 + len(ROWS)
    # Ýagdaý sütuny üçin sanaw (dropdown) – Garaşylýar / Tamamlandy
    dv = DataValidation(type="list", formula1='"Garaşylýar,Tamamlandy"', allow_blank=False)
    dv.error, dv.errorTitle = "Garaşylýar ýa-da Tamamlandy saýlaň", "Ýalňyş"
    ws.add_data_validation(dv)
    dv.add(f"F4:F{last_row}")

    # Sütun ini
    ws.column_dimensions["A"].width = 5   # №
    ws.column_dimensions["B"].width = 56  # Wepaly
    ws.column_dimensions["C"].width = 12  # Deadline
    ws.column_dimensions["D"].width = 10  # Bölüm
    ws.column_dimensions["E"].width = 10  # Dereje
    ws.column_dimensions["F"].width = 14  # Ýagdaý
    ws.column_dimensions["G"].width = 26  # Bellik

    # Auto-filter
    ws.auto_filter.ref = f"A3:G{last_row}"

    # Freeze: ilk 3 setir (başlyk + header)
    ws.freeze_panes = "A4"

    # ---- Sahypa 2: Özet (Deadline boýunça) ----
    ws2 = wb.create_sheet("Özet – Aý boýunça", 1)
    ws2["A1"] = "Deadline (aý) boýunça iş sany"
    ws2["A1"].font = Font(bold=True, size=12)
    ws2["A2"] = "Aý"
    ws2["B2"] = "Iş sany"
    for c in range(1, 3):
        ws2.cell(row=2, column=c).font = header_font
        ws2.cell(row=2, column=c).fill = header_fill
        ws2.cell(row=2, column=c).border = thin_border
    summary = {}
    for row in ROWS:
        date_str = row[1]  # YYYY-MM-DD
        month_key = date_str[:7]  # YYYY-MM
        summary[month_key] = summary.get(month_key, 0) + 1
    for idx, (month, count) in enumerate(sorted(summary.items()), 3):
        ws2.cell(row=idx, column=1, value=month)
        ws2.cell(row=idx, column=2, value=count)
        ws2.cell(row=idx, column=1).border = thin_border
        ws2.cell(row=idx, column=2).border = thin_border
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12

    # ---- Sahypa 3: Bölüm boýunça ----
    ws3 = wb.create_sheet("Özet – Bölüm", 2)
    ws3["A1"] = "Bölüm boýunça iş sany"
    ws3["A1"].font = Font(bold=True, size=12)
    ws3["A2"] = "Bölüm"
    ws3["B2"] = "Iş sany"
    for c in range(1, 3):
        ws3.cell(row=2, column=c).font = header_font
        ws3.cell(row=2, column=c).fill = header_fill
        ws3.cell(row=2, column=c).border = thin_border
    by_part = {}
    for row in ROWS:
        part = row[2]
        by_part[part] = by_part.get(part, 0) + 1
    for idx, (part, count) in enumerate(sorted(by_part.items()), 3):
        ws3.cell(row=idx, column=1, value=part)
        ws3.cell(row=idx, column=2, value=count)
        ws3.cell(row=idx, column=1).border = thin_border
        ws3.cell(row=idx, column=2).border = thin_border
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 12

    wb.save(OUTPUT_HOME)
    wb.save(OUTPUT_MAIN)
    print("Saved:", OUTPUT_HOME)
    print("Saved:", OUTPUT_MAIN)

if __name__ == "__main__":
    main()
